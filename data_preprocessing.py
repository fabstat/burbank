"""Burbank Preprocessor 

Preprocesses potato trials data and GDD data, and combines them.

This file contains a method to preprocess xlsx data files of a very specific
structure. See sample xlsx file in the xlsx_data folder.

It also contains a method to preprocess scraped GDD data from 
https://www.usbr.gov/pn/agrimet/webarcread.html
"""


import pandas as pd
import numpy as np
import openpyxl
import requests
import csv
from pathlib import Path
from glob import glob
import os
import re

#pd.set_option('future.no_silent_downcasting', True)


def scrape_website_data(station, year, start_month, start_day, end_month, end_day):
    # Dictionary mapping stations to alternative names
    station_mapping = {
        'HERO': 'herm_early' if start_month < 4 else 'herm_late',
        'KFLO': 'kf',
        'ONTO': 'ont',   
    }
    
    try:
        # Send GET request to the URL
        url = f"https://www.usbr.gov/pn-bin/daily.pl?station={station}&year={year}&month={start_month}&day={start_day}&year={year}&month={end_month}&day={end_day}&pcode=TG"
        response = requests.get(url)
        response.raise_for_status()
        
        # Extract content between BEGIN DATA and END DATA lines
        content = response.text
        data_match = re.search(r'BEGIN DATA\r?\n(.*?)\r?\nEND DATA', content, re.DOTALL)
        
        if not data_match:
            print("No data found between BEGIN DATA and END DATA")
            return
        
        # Extract and process the data
        raw_data = data_match.group(1).strip().split('\n')
        
        # Determine output filename
        year = re.search(r'year=(\d{4})', url).group(1)
        alt_station = station_mapping.get(station)
        output_filename = f'./gdd/{year}_{alt_station}.csv'
        
        # Write data to CSV
        with open(output_filename, 'w', newline='') as csvfile:
            csv_writer = csv.writer(csvfile)
            for line in raw_data:
                # Split the line and remove any extra whitespace
                fields = [field.strip() for field in line.split(',')]
                if len(fields) >= 2:
                    date = fields[0]
                    tg = fields[1]
                    csv_writer.writerow([date, tg])
        
        print(f"Data saved to {output_filename}")
    
    except requests.RequestException as e:
        print(f"Error fetching data: {e}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")




def extract_years_difference(year_range):
    """
    Convert a year range string to the number of years in trial
    
    Args:
        year_range (str): A string in the format "year1-year2"
    
    Returns:
        int: Number of years in trial (year2 - year1)
    """
    try:
        years = year_range.split('-')
        if len(years) == 2:
            rg = int(years[1]) - int(years[0])
            if rg > 3:
                return 0
            else:
                return rg
        else:
            return 1
    except (ValueError, IndexError):
        return 0
    
def sheet_adaptor(potato_obj, sheet, col, training):
    df1 = pd.DataFrame([potato_obj[sheet]['D1'].value], columns=['Clone'])
    if training:
        df2 = pd.DataFrame([potato_obj[sheet]['H61'].value], columns=['Keep'])
    df3 = pd.DataFrame([2024], columns=['Year'])
    df4 = pd.DataFrame([extract_years_difference(str(potato_obj[sheet]['B5'].value))], columns=['Yr in trial'])
    df5 = pd.DataFrame(potato_obj[sheet].iter_cols(min_row=11, min_col=col, max_col=col, max_row=12, values_only=True), 
                       columns=['Trial Region','Total Yield'])
    df6 = pd.DataFrame(potato_obj[sheet].iter_cols(min_row=14, min_col=col, max_col=col, max_row=15, values_only=True), 
                       columns=['%RB', 'Y1s'])
    df7 = pd.DataFrame(potato_obj[sheet].iter_cols(min_row=17, min_col=col, max_col=col, max_row=18, values_only=True), 
                       columns=['%RB 1s', '%1s'])
    df8 = pd.DataFrame(potato_obj[sheet].iter_cols(min_row=22, min_col=col, max_col=col, max_row=22, values_only=True), 
                       columns=['Over 20oz'])
    df9 = pd.DataFrame(potato_obj[sheet].iter_cols(min_row=19, min_col=col, max_col=col, max_row=21, values_only=True), 
                       columns=['10-20oz','6-10oz','4-6oz'])
    df10 = pd.DataFrame(potato_obj[sheet].iter_cols(min_row=24, min_col=col, max_col=col, max_row=24, values_only=True), 
                       columns=['Under 4oz'])
    df11 = pd.DataFrame(potato_obj[sheet].iter_cols(min_row=23, min_col=col, max_col=col, max_row=23, values_only=True), 
                       columns=['Y2s'])
    df12 = pd.DataFrame(potato_obj[sheet].iter_cols(min_row=25, min_col=col, max_col=col, max_row=25, values_only=True), 
                       columns=['Culls'])
    df13 = pd.DataFrame(potato_obj[sheet].iter_cols(min_row=26, min_col=col, max_col=col, max_row=29, values_only=True), 
                       columns=['Tuber/Plant', 'Ave. Tuber Size', 'Length/Width', 'Specific Gravity'])
    df14 = pd.DataFrame(potato_obj[sheet].iter_cols(min_row=30, min_col=col, max_col=col, max_row=32, values_only=True), 
                       columns=['Fry Color Stem', 'Fry Color Bud', 'Sugar Ends'])
    df15 = pd.DataFrame(potato_obj[sheet].iter_cols(min_row=33, min_col=col, max_col=col, max_row=36, values_only=True), 
                       columns=['Hollow Heart', 'Brown Center', 'IBS', 'Blackspot'])
    df16 = pd.DataFrame(potato_obj[sheet].iter_cols(min_row=37, min_col=col, max_col=col, max_row=39, values_only=True), 
                       columns=['Vascular', 'Skin Color', 'Russeting'])
    df17 = pd.DataFrame(potato_obj[sheet].iter_cols(min_row=41, min_col=col, max_col=col, max_row=42, values_only=True), 
                       columns=['Tuber Shape','Shape Uniformity'])
    df18 = pd.DataFrame(potato_obj[sheet].iter_cols(min_row=40, min_col=col, max_col=col, max_row=40, values_only=True), 
                       columns=['Eye Depth'])
    df19 = pd.DataFrame(potato_obj[sheet].iter_cols(min_row=47, min_col=col, max_col=col, max_row=47, values_only=True), 
                       columns=['Greening'])
    df20 = pd.DataFrame(potato_obj[sheet].iter_cols(min_row=43, min_col=col, max_col=col, max_row=45, values_only=True), 
                       columns=['Growth Cracks', 'Scab', 'Shatter Bruise'])
    if training:
        df21 = pd.concat([df1, df2, df3, df4, df5, df6, df7, df8, df9, df10, 
                          df11, df12, df13, df14, df15, df16, df17, df18, df19, df20], axis=1)   
    else: 
        df21 = pd.concat([df1, df3, df4, df5, df6, df7, df8, df9, df10, 
                          df11, df12, df13, df14, df15, df16, df17, df18, df19, df20], axis=1) 
    #df = pd.concat([df, df21], axis=0)
    df21['Y2s + Over 20oz'] = df21['Y2s']+df21['Over 20oz']
    return df21



def process_xlsx_data(path_to_xlsx_files, year, training=False):
    """Processes xlsx files of a particular structure
    
    Parameters
    ----------
    path_to_xlsx_files: str
        The path to the xlsx files
    training: boolean
        If true, input data must contain a 'Keep' variable
        otherwise the data is considered for prediction only
    """
    combined_df = pd.DataFrame()
    for potato_file in Path(path_to_xlsx_files).glob(f"*{year}.xlsx"):
        potato_obj = openpyxl.load_workbook(potato_file, data_only=True)
        df = pd.DataFrame() 
        for sheet in potato_obj.sheetnames: 
            if sheet in ['AOR161181', 'AOR180691']:
                range_start = 4
                range_end = 7        
            elif sheet in ['Shepody']:
                range_start = 4
                range_end = 5
            else:
                range_start = 4
                range_end = 8
                
            for col in range(range_start, range_end):
                df21 = sheet_adaptor(potato_obj, sheet, col, training)
                df = pd.concat([df, df21], axis=0)
        combined_df = pd.concat([combined_df, df], axis=0)

    
    combined_df['Year'] = combined_df['Year'].astype(int)
    # combined_df['Yr in trial'] = combined_df['Yr in trial'].mask((combined_df['Yr in trial']=='Ctrl') | 
    #                                                              (combined_df['Yr in trial']=='Control'), other='0').astype('int')
    combined_df['Trial Region'] = combined_df['Trial Region'].str.replace('\n', ' ')
    combined_df = combined_df.replace(["Ontario", "K Falls", "Hermiston Early State", "Hermiston Late State"], ["ONT", "KF", "HER_Early", "HER_Late"])


    if training:
        combined_df.loc[combined_df['Keep'] == 'keep', 'Keep'] = 1
        combined_df.loc[combined_df['Keep'] == 'drop', 'Keep'] = 0
        # For each clone, the max number of years in trial
        dur_tb = combined_df.groupby(by=['Clone']).agg({'Yr in trial': 'max',
                                                        'Keep': 'min',
                                                        'Year': 'max'}).sort_values(by='Yr in trial', ascending=False)
        # Potatoes graduate if they are not dropped for three years
        true_keeps = list(dur_tb[(dur_tb['Yr in trial'] >= 3) & (dur_tb['Keep'] == 1)].index)
        # Potatoes currently on trial
        # maybe_keeps = list(dur_tb[(dur_tb['Yr in trial'] <= 2) & 
        #                           (dur_tb['Year'] == 2021) & 
        #                           (dur_tb['Keep'] == 1)].index)
        # Potatoes survived 2 years in trial, but were dropped in third year
        # almost_keeps = list(dur_tb[
        #                    ((dur_tb['Yr in trial'] <= 2) & (dur_tb['Keep'] == 1) & (dur_tb['Year'] != 2021)) |
        #                    ((dur_tb['Yr in trial'] == 3) & (dur_tb['Keep'] == 0))].index)
        # Potatoes survived 1 year in trial, but were dropped in next year
        true_drops = list(dur_tb[(dur_tb['Yr in trial'] <= 3) & (dur_tb['Keep'] == 0)].index)
        
        # Season's control - check gdd relationship
        ctrl = list(dur_tb[dur_tb['Yr in trial'] == 0].index)
        
        combined_df['true_keeps'] = None
        #combined_df.loc[combined_df['Clone'].isin(maybe_keeps), 'true_keeps'] = 2
        combined_df.loc[combined_df['Clone'].isin(true_keeps), 'true_keeps'] = 1
        combined_df.loc[combined_df['Clone'].isin(true_drops), 'true_keeps'] = 0
        #combined_df.loc[combined_df['Clone'].isin(almost_keeps), 'true_keeps'] = -2
        combined_df.loc[combined_df['Clone'].isin(ctrl), 'true_keeps'] = -1
        combined_df['true_keeps'] = combined_df['true_keeps'].astype("float")
        
    combined_df['Ctrl ave'] = None

    ##### Places and years: must change these manually if needed ####
    places = ["ONT", "KF", "HER_Early", "HER_Late"] #['ONT', 'HER', 'HER_Early', 'HER_Late', 'COR', 'KF']
    years = [2024] #list(range(2013,2022))
    #################################################################
    for year in years:
        for place in places:
            combined_df.loc[(combined_df['Year']==year) & (combined_df['Trial Region']==place), 'Ctrl ave'] = combined_df.loc[
                (combined_df['Year']==year) & (combined_df['Trial Region']==place) & (combined_df['Yr in trial'] == 0), 'Total Yield'].mean()

    combined_df['% CA'] = 100*combined_df['Total Yield']/combined_df['Ctrl ave']
    combined_df[['Ctrl ave', '% CA']] = combined_df[['Ctrl ave', '% CA']].astype("float")
        
    return combined_df # potatoes.csv



def make_gdd_dict(): 
    # Data comes from: https://www.usbr.gov/pn/agrimet/webarcread.html
    locations = ['herm_early', 'herm_late', 'kf', 'ont']
    years = [2024]
    gddict = pd.DataFrame(columns = ['Year', 'Trial Region', 'GDD 1-60', 'GDD 61-90', 'GDD 91-end'])

    for year in years:
        for location in locations:
            try:
                df = pd.read_csv(f'./gdd/{year:d}_{location:s}.csv', na_values='NA')
            except:
                continue

            first = df[df.columns[1]].iloc[0:60].sum()
            second = df[df.columns[1]].iloc[60:90].sum()
            third = df[df.columns[1]].iloc[90:].sum()

            new_row = pd.Series({'Year': year, 
                                 'Trial Region': 
                                 location[:3].upper() + '_' + location[5:].capitalize() 
                                 if '_' in location 
                                 else location.upper(),
                                 'GDD 1-60': first, 
                                 'GDD 61-90': second, 
                                 'GDD 91-end': third})

            gddict = pd.concat([gddict, new_row.to_frame().T], ignore_index=True) 
    
    return gddict # gdd.csv


def load_preprocessed_data(path_to_data):
    df = pd.read_csv(f"{path_to_data}/potatoes.csv") 
    gdds = pd.read_csv(f"{path_to_data}/gdd.csv")
    gddf = df.set_index(pd.MultiIndex.from_frame(df[['Year', 'Trial Region']])).join(
        gdds.set_index(pd.MultiIndex.from_frame(gdds[['Year', 'Trial Region']])), 
        rsuffix='_a').reset_index(drop=True).drop(['Unnamed: 0', 'Year_a', 'Trial Region_a'], axis=1)
 
    return gddf